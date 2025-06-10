from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, String
from typing import List, Optional

from app.database import get_db
from app.models.user import User
from app.models.queue import QueueEntry, QueueStatus
from app.models.video import VideoSettings
from app.schemas.queue import QueueResponse
from app.schemas.video import VideoSettingsResponse, VideoSettingsUpdate
from app.schemas import AdminUserCreate, UserResponse, UserUpdate
from app.security import get_admin_user
from app.services.user import create_user
from app.services.queue import get_all_queue_entries
from app.services.archive import get_archive_statistics, cleanup_old_completed_entries
from app.models.archive import ArchivedQueueEntry
from fastapi.responses import StreamingResponse
import io
import csv
from openpyxl import Workbook 
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

# Полный словарь для маппинга названий программ на коды
PROGRAM_MAPPING = {
    # БАКАЛАВРИАТ
    # Русские названия
    "бухгалтерский учёт": "accounting",
    "бухгалтерский учет": "accounting",
    "прикладная лингвистика": "appliedLinguistics",
    "экономика и наука о данных": "economicsDataScience",
    "финансы": "finance",
    "гостеприимство": "hospitality",
    "международная журналистика": "internationalJournalism",
    "международное право": "internationalLaw",
    "международные отношения": "internationalRelations",
    "it": "it",
    "ит": "it",
    "юриспруденция": "jurisprudence",
    "менеджмент": "management",
    "маркетинг": "marketing",
    "психология": "psychology",
    "туризм": "tourism",
    "переводческое дело": "translation",
    
    # Казахские названия
    "бухгалтерлік есеп": "accounting",
    "қолданбалы лингвистика": "appliedLinguistics",
    "экономика және деректер ғылымы": "economicsDataScience",
    "қаржы": "finance",
    "қонақжайлылық": "hospitality",
    "халықаралық журналистика": "internationalJournalism",
    "халықаралық құқық": "internationalLaw",
    "халықаралық қатынастар": "internationalRelations",
    "құқықтану": "jurisprudence",
    "аударма ісі": "translation",
    
    # Английские названия
    "accounting": "accounting",
    "applied linguistics": "appliedLinguistics",
    "economics and data science": "economicsDataScience",
    "finance": "finance",
    "hospitality": "hospitality",
    "international journalism": "internationalJournalism",
    "international law": "internationalLaw",
    "international relations": "internationalRelations",
    "law": "jurisprudence",
    "management": "management",
    "marketing": "marketing",
    "psychology": "psychology",
    "tourism": "tourism",
    "translation studies": "translation",
    
    # МАГИСТРАТУРА
    # Русские названия
    "политология и международные отношения": "politicalInternationalRelations",
    "конкурентное право": "competitionLaw",
    "консультативная психология": "consultingPsychology",
    "экономика": "economics",
    "право интеллектуальной собственности и бизнеса": "intellectualPropertyLaw",
    "право it": "itLaw",
    "право ит": "itLaw",
    
    # Казахские названия
    "саясаттану және халықаралық қатынастар": "politicalInternationalRelations",
    "бәсекелестік құқық": "competitionLaw",
    "консультативті психология": "consultingPsychology",
    "зияткерлік меншік және бизнес құқық": "intellectualPropertyLaw",
    "құқық it": "itLaw",
    
    # Английские названия
    "political science and international relations": "politicalInternationalRelations",
    "competition law": "competitionLaw",
    "counselling psychology": "consultingPsychology",
    "economics": "economics",
    "intellectual property and business law": "intellectualPropertyLaw",
    "it law": "itLaw",
    
    # ДОКТОРАНТУРА
    # Русские названия
    "право": "law",
    "phd по экономике": "phdEconomics",
    
    # Казахские названия
    "құқық": "law",
    "экономика саласындағы phd": "phdEconomics",
    
    # Английские названия
    "phd in law": "law",
    "phd in economics": "phdEconomics"
}

def get_program_codes_by_name(program_name: str) -> List[str]:
    """
    Получает возможные коды программ по названию программы
    """
    if not program_name:
        return []
    
    program_name_lower = program_name.lower().strip()
    matching_codes = []
    
    # Ищем точные совпадения
    if program_name_lower in PROGRAM_MAPPING:
        matching_codes.append(PROGRAM_MAPPING[program_name_lower])
    
    # Ищем частичные совпадения в названиях
    for name, code in PROGRAM_MAPPING.items():
        if program_name_lower in name or name in program_name_lower:
            if code not in matching_codes:
                matching_codes.append(code)
    
    # Если ничего не найдено, возможно пользователь ввел код напрямую
    if not matching_codes:
        matching_codes.append(program_name_lower)
    
    return matching_codes

@router.get("/queue/export")
def export_queue_to_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Export all queue entries to Excel (.xlsx)"""
    # Get all queue entries
    queue_entries = get_all_queue_entries(db)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Queue Data"
    
    # Headers
    headers = ["ФИО", "Программы", "Номер", "Сотрудник", "Дата создания", "Статус", "Время обработки (сек)"]
    
    # Add headers to worksheet
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # Style headers
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row, entry in enumerate(queue_entries, 2):
        programs = ", ".join(entry.programs) if isinstance(entry.programs, list) else entry.programs
        created_at = entry.created_at.strftime("%Y-%m-%d %H:%M:%S") if entry.created_at else "-"
        
        data_row = [
            entry.full_name,
            programs,
            entry.queue_number,
            entry.assigned_employee_name or "-",
            created_at,
            entry.status.value,
            entry.processing_time or "-"
        ]
        
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return file as response
    headers = {
        'Content-Disposition': 'attachment; filename="queue_data.xlsx"'
    }
    return StreamingResponse(
        io.BytesIO(output.getvalue()), 
        headers=headers, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/queue/reset-numbering")
def reset_queue_numbering(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Сбросить нумерацию очереди (только для админов)"""
    try:
        # Получаем все активные заявки (не completed)
        active_entries = db.query(QueueEntry).filter(
            QueueEntry.status.in_([QueueStatus.WAITING, QueueStatus.IN_PROGRESS, QueueStatus.PAUSED])
        ).order_by(QueueEntry.created_at).all()
        
        # Архивируем и удаляем все completed заявки
        completed_entries = db.query(QueueEntry).filter(QueueEntry.status == QueueStatus.COMPLETED).all()
        
        archived_count = 0
        for entry in completed_entries:
            try:
                archive_queue_entry(db, entry, reason="manual_reset")
                db.delete(entry)
                archived_count += 1
            except Exception as e:
                continue
        
        # Перенумеровываем активные заявки начиная с 1
        renumbered_count = 0
        for i, entry in enumerate(active_entries, 1):
            entry.queue_number = i
            db.add(entry)
            renumbered_count += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Queue numbering reset successfully",
            "archived_completed": archived_count,
            "renumbered_active": renumbered_count,
            "next_number": len(active_entries) + 1
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Reset failed: {str(e)}")

@router.post("/create-admission", response_model=UserResponse)
def create_admission_staff(
    user_data: AdminUserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Create a new admission staff member (admin only)"""
    # Create a new user with admission role
    return create_user(db=db, user=user_data, role="admission")

@router.get("/employees", response_model=List[UserResponse])
def get_all_employees(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Get all employees (admin only)"""
    employees = db.query(User).filter(User.role == "admission").all()
    return employees

@router.get("/queue", response_model=List[QueueResponse])
def get_all_queue_entries_api(
    status: Optional[QueueStatus] = None,
    date: Optional[str] = None,
    employee: Optional[str] = None,
    full_name: Optional[str] = None,
    program: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Get all queue entries with filters (admin only)"""
    from datetime import datetime
    from sqlalchemy import func, and_, or_, text
    
    # Начинаем с базового запроса
    query = db.query(QueueEntry)
    
    # Применяем фильтры
    if status:
        query = query.filter(QueueEntry.status == status)
    
    if date:
        # Фильтруем по дате (формат YYYY-MM-DD)
        try:
            filter_date = datetime.strptime(date, "%Y-%m-%d")
            # Фильтруем записи созданные в этот день
            start_of_day = filter_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_of_day = filter_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            query = query.filter(
                and_(
                    QueueEntry.created_at >= start_of_day,
                    QueueEntry.created_at <= end_of_day
                )
            )
        except ValueError:
            # Если неправильный формат даты, игнорируем фильтр
            pass
    
    if employee:
        # Фильтруем по имени сотрудника (частичное совпадение, без учета регистра)
        query = query.filter(
            QueueEntry.assigned_employee_name.ilike(f"%{employee}%")
        )
    
    if full_name:
        # Фильтруем по ФИО абитуриента (частичное совпадение, без учета регистра)
        query = query.filter(
            QueueEntry.full_name.ilike(f"%{full_name}%")
        )
    
    if program:
        # Получаем возможные коды программ по введенному названию
        program_codes = get_program_codes_by_name(program)
        
        if program_codes:
            # Создаем условия для поиска по всем возможным кодам
            program_conditions = []
            for code in program_codes:
                # Ищем код в JSON массиве программ
                program_conditions.append(
                    text("programs::text ILIKE :code").params(code=f'%"{code}"%')
                )
            
            # Объединяем условия через OR
            if len(program_conditions) == 1:
                query = query.filter(program_conditions[0])
            else:
                query = query.filter(or_(*program_conditions))
    
    return query.all()

@router.delete("/employees/{user_id}")
def delete_employee(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Delete employee (admin only)"""
    employee = db.query(User).filter(User.id == user_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    db.delete(employee)
    db.commit()
    return {"detail": "Employee deleted successfully"}

@router.put("/employees/{user_id}", response_model=UserResponse)
def update_employee(
    user_id: str,
    user_data: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Update employee data (admin only)"""
    employee = db.query(User).filter(User.id == user_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    for key, value in user_data.dict(exclude_unset=True).items():
        setattr(employee, key, value)
    
    db.commit()
    db.refresh(employee)
    return employee

# === РОУТЫ ДЛЯ УПРАВЛЕНИЯ ВИДЕО ===

@router.get("/video-settings", response_model=VideoSettingsResponse)
def get_video_settings(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Get current video settings (admin only)"""
    settings = db.query(VideoSettings).first()
    if not settings:
        # Создаем запись если её нет
        settings = VideoSettings(youtube_url="", is_enabled=False)
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings

@router.put("/video-settings", response_model=VideoSettingsResponse)
def update_video_settings(
    video_data: VideoSettingsUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    """Update video settings (admin only)"""
    settings = db.query(VideoSettings).first()
    if not settings:
        # Создаем новую запись если её нет
        settings = VideoSettings()
        db.add(settings)
    
    # Обновляем поля
    for key, value in video_data.dict(exclude_unset=True).items():
        setattr(settings, key, value)
    
    db.commit()
    db.refresh(settings)
    return settings